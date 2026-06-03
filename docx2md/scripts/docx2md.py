"""
使用pandoc将多种文档格式转换为markdown，提取图片和嵌入附件

支持格式：
- docx: Word文档，提取图片和OLE附件
- xlsx/xls: Excel文件，转换为markdown表格
"""
import subprocess
import os
import re
import shutil
import zipfile
import struct
import io
import tempfile
from pathlib import Path
from PIL import Image
from openpyxl import load_workbook
import pandas as pd
import xlrd


def get_ole10native_filename(data):
    """从Package对象获取原始文件名和嵌入数据"""
    if data[:8] != b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return None, None

    sector_size = 512
    dir_start = struct.unpack('<I', data[48:52])[0]
    dir_offset = 512 + dir_start * sector_size

    for i in range(16):
        entry_offset = dir_offset + i * 128
        if entry_offset + 128 > len(data):
            break

        entry_data = data[entry_offset:entry_offset + 128]
        name_len = struct.unpack('<H', entry_data[64:66])[0]
        if name_len > 0 and name_len <= 64:
            name_bytes = entry_data[:name_len]
            try:
                name = name_bytes.decode('utf-16-le', errors='replace').rstrip('\x00')
            except:
                continue

            if name == '\x01Ole10Native':
                start_sector = struct.unpack('<I', entry_data[116:120])[0]
                stream_size = struct.unpack('<I', entry_data[120:124])[0]

                stream_offset = 512 + start_sector * sector_size
                if stream_offset < len(data):
                    stream_data = data[stream_offset:stream_offset + stream_size]

                    # Package Ole10Native格式:
                    # [4 bytes] 数据大小
                    # [2 bytes] flags
                    # [GBK字符串] 原始文件名，null结尾
                    # [ANSI字符串] 临时文件路径，null结尾
                    # [额外元数据] 可能包含另一个路径
                    # [文件数据] 实际嵌入内容

                    # 从位置6开始找文件名null结束
                    filename_start = 6
                    filename_end = stream_data.find(b'\x00', filename_start)
                    if filename_end != -1 and filename_end < 300:
                        filename_bytes = stream_data[filename_start:filename_end]
                        filename = filename_bytes.decode('gbk', errors='replace')

                        # 找临时路径结束
                        temp_path_end = stream_data.find(b'\x00', filename_end + 1)

                        # 在数据中搜索文件类型签名，获取实际内容
                        # PNG签名
                        png_pos = stream_data.find(b'\x89PNG\r\n\x1a\n', temp_path_end)
                        if png_pos != -1:
                            embedded_data = stream_data[png_pos:]
                            # 找PNG结束(IEND)
                            iend_pos = embedded_data.find(b'IEND')
                            if iend_pos != -1:
                                embedded_data = embedded_data[:iend_pos + 12]
                            return filename, embedded_data

                        # BMP签名
                        bmp_pos = stream_data.find(b'BM', temp_path_end)
                        if bmp_pos != -1:
                            # BMP数据需要从BM开始到文件结束
                            # BMP文件头包含文件大小信息
                            bmp_header = stream_data[bmp_pos:bmp_pos + 54]
                            if len(bmp_header) >= 54:
                                bmp_file_size = struct.unpack('<I', bmp_header[2:6])[0]
                                embedded_data = stream_data[bmp_pos:bmp_pos + bmp_file_size]
                                return filename, embedded_data

                        # JPEG签名
                        jpeg_pos = stream_data.find(b'\xff\xd8\xff', temp_path_end)
                        if jpeg_pos != -1:
                            # 找JPEG结束标记
                            jpeg_end = stream_data.find(b'\xff\xd9', jpeg_pos)
                            if jpeg_end != -1:
                                embedded_data = stream_data[jpeg_pos:jpeg_end + 2]
                                return filename, embedded_data

                        # 如果没找到已知签名，从temp_path_end+1开始作为数据
                        embedded_data = stream_data[temp_path_end + 1:]
                        return filename, embedded_data
    return None, None


def extract_xlsx_from_ole(data):
    """从Excel OLE对象提取xlsx - 改进版本"""
    # xlsx是zip格式，必须找到完整的zip结构
    # 在OLE对象中，xlsx数据通常从特定位置开始

    # 找到第一个PK (local file header) - 通常在固定偏移后
    pk_local = data.find(b'PK\x03\x04')
    if pk_local == -1:
        return None

    # 找到EOCD (End of Central Directory)
    pk_eocd = data.find(b'PK\x05\x06')
    if pk_eocd == -1:
        return None

    # 读取EOCD获取注释长度
    eocd_data = data[pk_eocd:pk_eocd + 22]
    comment_len = struct.unpack('<H', eocd_data[20:22])[0]
    eocd_end = pk_eocd + 22 + comment_len

    # 提取完整zip数据
    xlsx_data = data[pk_local:eocd_end]

    # 验证zip完整性
    try:
        z = zipfile.ZipFile(io.BytesIO(xlsx_data))
        # 检查关键文件是否存在
        if '[Content_Types].xml' not in z.namelist():
            return None
        return xlsx_data
    except:
        return None


def _convert_docx_bytes_to_md(data: bytes, parent_images_dir=None, parent_attachments_dir=None, next_img_idx=1):
    """将docx字节数据转为markdown，支持递归提取嵌套OLE附件

    Args:
        data: docx文件字节数据
        parent_images_dir: 父级images目录（用于合并子文档图片）
        parent_attachments_dir: 父级attachments目录（用于合并子文档附件）
        next_img_idx: 图片起始编号

    Returns:
        (md_text, next_img_idx): markdown文本和下一个可用图片编号
    """
    import tempfile
    tmp_dir = tempfile.mkdtemp(prefix="docx2md_nested_")
    try:
        tmp_docx = os.path.join(tmp_dir, "att.docx")
        with open(tmp_docx, 'wb') as f:
            f.write(data)

        # 走完整的docx_to_md流程（含OLE提取、图片处理）
        md_path = docx_to_md(tmp_docx, tmp_dir)
        md_text = Path(md_path).read_text(encoding='utf-8')

        stem = Path(tmp_docx).stem
        child_images_dir = Path(tmp_dir) / f"{stem}_files" / "images"
        child_attachments_dir = Path(tmp_dir) / f"{stem}_files" / "attachments"

        # 合并图片到父级目录
        if parent_images_dir and child_images_dir.exists():
            parent_images_dir.mkdir(parents=True, exist_ok=True)
            for img_file in sorted(child_images_dir.iterdir()):
                if img_file.is_file():
                    new_name = f"image_{next_img_idx}{img_file.suffix}"
                    shutil.copy2(img_file, parent_images_dir / new_name)
                    md_text = md_text.replace(
                        f"![]({stem}_files/images/{img_file.name})",
                        f"![]({Path(parent_images_dir).parent.name}/images/{new_name})"
                    )
                    next_img_idx += 1

        # 合并附件到父级目录
        if parent_attachments_dir and child_attachments_dir.exists():
            parent_attachments_dir.mkdir(parents=True, exist_ok=True)
            for att_file in child_attachments_dir.iterdir():
                if att_file.is_file():
                    shutil.copy2(att_file, parent_attachments_dir / att_file.name)

        # 修正嵌套文档中的双链路径：att_files/attachments/ → 父级实际路径
        if parent_attachments_dir:
            parent_prefix = f"{Path(parent_attachments_dir).parent.name}/attachments/"
            md_text = md_text.replace(f"{stem}_files/attachments/", parent_prefix)

        return md_text, next_img_idx
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _convert_doc_bytes_to_md(data: bytes, original_name: str) -> str:
    """将doc字节数据转为markdown：先用Word COM转docx，再用pandoc转md"""
    import tempfile
    tmp_dir = tempfile.mkdtemp(prefix="docx2md_att_")
    tmp_doc = str(Path(tmp_dir) / original_name)
    tmp_docx = str(Path(tmp_dir) / (Path(original_name).stem + ".docx"))
    try:
        with open(tmp_doc, 'wb') as f:
            f.write(data)

        try:
            import win32com.client
        except ImportError:
            raise RuntimeError("转换.doc附件需要pywin32库")

        word = win32com.client.Dispatch("Word.Application")
        word.Visible = False
        try:
            doc = word.Documents.Open(tmp_doc)
            doc.SaveAs(tmp_docx, FileFormat=16)
            doc.Close()
        finally:
            word.Quit()

        return _convert_docx_bytes_to_md(open(tmp_docx, 'rb').read())
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def extract_ole_objects(docx_path, assets_dir, base_name):
    """从docx中提取OLE嵌入对象（按需创建目录）"""
    ole_info = {}  # image_name -> {'original_name': str, 'saved_name': str, 'is_image': bool, 'image_idx': int}
    image_idx = 1  # 图片编号计数器

    # 目录路径（按需创建）
    images_dir = assets_dir / "images"
    attachments_dir = assets_dir / "attachments"

    with zipfile.ZipFile(docx_path, 'r') as z:
        # 读取关系文件
        rels_xml = z.read('word/_rels/document.xml.rels').decode('utf-8')

        # 解析所有关系
        rels = {}
        for match in re.finditer(r'<Relationship Id="rId(\d+)"[^>]*Type="[^"]*relationships/([^"]+)"[^>]*Target="([^"]+)"', rels_xml):
            rId = int(match.group(1))
            rel_type = match.group(2)
            target = match.group(3)
            rels[rId] = (rel_type, target)

        # 找OLE对象和对应图片的映射
        ole_to_image = {}
        for rId, (rel_type, target) in sorted(rels.items()):
            if rel_type == 'oleObject' and 'embeddings/' in target:
                ole_num_match = re.search(r'oleObject(\d+)', target)
                if ole_num_match:
                    ole_num = int(ole_num_match.group(1))
                    next_rId = rId + 1
                    if next_rId in rels:
                        next_type, next_target = rels[next_rId]
                        if next_type == 'image' and 'media/' in next_target:
                            image_name = next_target.split('/')[-1]
                            ole_to_image[ole_num] = image_name

        # 读取document.xml获取OLE对象的ProgID
        doc_xml = z.read('word/document.xml').decode('utf-8')

        ole_progids = {}
        for match in re.finditer(r'<o:OLEObject[^>]*ProgID="([^"]+)"[^>]*r:id="rId(\d+)"', doc_xml):
            prog_id = match.group(1)
            ole_rId = int(match.group(2))
            ole_progids[ole_rId] = prog_id

        # 处理每个OLE对象
        for ole_rId, prog_id in ole_progids.items():
            if ole_rId not in rels:
                continue
            _, ole_target = rels[ole_rId]
            ole_num_match = re.search(r'oleObject(\d+)', ole_target)
            if not ole_num_match:
                continue
            ole_num = int(ole_num_match.group(1))

            ole_path = f'word/{ole_target}'
            if ole_path not in z.namelist():
                continue

            ole_data = z.read(ole_path)
            image_name = ole_to_image.get(ole_num)

            original_name = None
            embedded_data = None
            is_image = False

            if prog_id == 'Package':
                original_name, embedded_data = get_ole10native_filename(ole_data)
                if original_name:
                    name_lower = original_name.lower()
                    # 检查是否是图片文件
                    if name_lower.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                        is_image = True
                    # 检查是否是Excel文件，转为md
                    elif name_lower.endswith(('.xlsx', '.xls')) and embedded_data:
                        try:
                            if name_lower.endswith('.xls'):
                                # 旧版.xls格式，用xlrd
                                wb_source = xlrd.open_workbook(file_contents=embedded_data)
                                sheet_names = wb_source.sheet_names()
                                md_content = ""
                                for sheet_name in sheet_names:
                                    ws = wb_source.sheet_by_name(sheet_name)
                                    md_content += f"## {sheet_name}\n\n"
                                    md_content += _xlrd_sheet_to_md(ws)
                                    md_content += "\n"
                            else:
                                # 新版.xlsx格式，用openpyxl
                                wb_source = load_workbook(io.BytesIO(embedded_data), data_only=True)
                                sheet_names = wb_source.sheetnames
                                md_content = ""
                                for sheet_name in sheet_names:
                                    ws = wb_source[sheet_name]
                                    md_content += f"## {sheet_name}\n\n"
                                    md_content += _worksheet_to_md(ws)
                                    md_content += "\n"
                                wb_source.close()

                            md_name = f'{Path(original_name).stem}.md'
                            safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", original_name)

                            attachments_dir.mkdir(parents=True, exist_ok=True)
                            att_path = attachments_dir / safe_name
                            att_path.write_bytes(embedded_data)
                            md_path = attachments_dir / md_name
                            md_path.write_text(md_content, encoding='utf-8')

                            print(f"提取附件(xlsx): {original_name}")
                            print(f"生成文本(md): {md_name}")

                            ole_info[image_name] = {
                                'original_name': original_name,
                                'saved_name': safe_name,
                                'md_name': md_name,
                                'is_image': False,
                                'has_md': True
                            }
                            continue
                        except Exception as e:
                            print(f"附件Excel处理失败({original_name}): {e}")
                            is_image = False
                    # 检查是否是Word文件，转为md
                    elif name_lower.endswith('.docx') and embedded_data:
                        try:
                            md_content, image_idx = _convert_docx_bytes_to_md(
                                embedded_data, images_dir, attachments_dir, image_idx
                            )
                            md_name = f'{Path(original_name).stem}.md'
                            safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", original_name)

                            attachments_dir.mkdir(parents=True, exist_ok=True)
                            att_path = attachments_dir / safe_name
                            att_path.write_bytes(embedded_data)
                            md_path_att = attachments_dir / md_name
                            md_path_att.write_text(md_content, encoding='utf-8')

                            print(f"提取附件(docx): {original_name}")
                            print(f"生成文本(md): {md_name}")

                            ole_info[image_name] = {
                                'original_name': original_name,
                                'saved_name': safe_name,
                                'md_name': md_name,
                                'is_image': False,
                                'has_md': True
                            }
                            continue
                        except Exception as e:
                            print(f"附件docx转md失败({original_name}): {e}")
                            is_image = False
                    elif name_lower.endswith('.doc') and embedded_data:
                        try:
                            md_content = _convert_doc_bytes_to_md(embedded_data, original_name)
                            md_name = f'{Path(original_name).stem}.md'
                            safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", original_name)

                            attachments_dir.mkdir(parents=True, exist_ok=True)
                            att_path = attachments_dir / safe_name
                            att_path.write_bytes(embedded_data)
                            md_path_att = attachments_dir / md_name
                            md_path_att.write_text(md_content, encoding='utf-8')

                            print(f"提取附件(doc): {original_name}")
                            print(f"生成文本(md): {md_name}")

                            ole_info[image_name] = {
                                'original_name': original_name,
                                'saved_name': safe_name,
                                'md_name': md_name,
                                'is_image': False,
                                'has_md': True
                            }
                            continue
                        except Exception as e:
                            print(f"附件doc转md失败({original_name}): {e}")
                            is_image = False
                    else:
                        is_image = False
            elif prog_id == 'Excel.Sheet.12':
                xlsx_data = extract_xlsx_from_ole(ole_data)
                if xlsx_data:
                    try:
                        # 用openpyxl读取数据
                        wb_source = load_workbook(io.BytesIO(xlsx_data), data_only=True)

                        # 获取工作表名称作为文件名
                        sheet_names = wb_source.sheetnames
                        if sheet_names:
                            original_name = sheet_names[0] + '.xlsx'
                        else:
                            original_name = f'附件{ole_num}.xlsx'

                        # 用pandas+xlsxwriter重新生成xlsx文件
                        new_buffer = io.BytesIO()
                        with pd.ExcelWriter(new_buffer, engine='xlsxwriter') as writer:
                            for sheet_name in sheet_names:
                                ws = wb_source[sheet_name]
                                data = []
                                for row in ws.iter_rows(values_only=True):
                                    data.append(list(row))
                                df = pd.DataFrame(data)
                                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                        embedded_data = new_buffer.getvalue()
                        is_image = False

                        # 同时生成md格式的文本文件
                        md_content = ""
                        for sheet_name in sheet_names:
                            ws = wb_source[sheet_name]
                            md_content += f"## {sheet_name}\n\n"
                            md_content += _worksheet_to_md(ws)
                            md_content += "\n"

                        wb_source.close()

                        # md文件名（与xlsx同名，扩展名改为.md）
                        md_name = original_name.replace('.xlsx', '.md')

                        print(f"Excel重新生成(xlsxwriter): {original_name}")
                        print(f"同时生成MD文本: {md_name}")

                        # 保存md文件（稍后在统一位置处理）
                        # 这里先记录md内容，在后面保存xlsx时同时保存md
                        ole_info[image_name] = {
                            'xlsx_name': original_name,
                            'xlsx_data': embedded_data,
                            'md_name': md_name,
                            'md_content': md_content,
                            'is_image': False
                        }
                        # 跳过后续的统一保存逻辑
                        continue
                    except Exception as e:
                        print(f"Excel处理失败(ole{ole_num}): {e}")
                        continue
            elif prog_id == 'Word.Document.12':
                # Word docx嵌入对象：OLE Compound Document中的"package"流就是docx
                docx_data = None
                if ole_data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                    # OLE Compound Document，提取package流
                    sector_size = 512
                    dir_start = struct.unpack('<I', ole_data[48:52])[0]
                    dir_offset = sector_size + dir_start * sector_size
                    for _i in range(10):
                        entry_offset = dir_offset + _i * 128
                        if entry_offset + 128 > len(ole_data):
                            break
                        entry = ole_data[entry_offset:entry_offset + 128]
                        name_len = struct.unpack('<H', entry[64:66])[0]
                        if name_len > 2:
                            name = entry[:name_len-2].decode('utf-16-le', errors='replace')
                            if name == 'package':
                                pkg_start = struct.unpack('<I', entry[116:120])[0]
                                pkg_size = struct.unpack('<I', entry[120:124])[0]
                                pkg_offset = sector_size + pkg_start * sector_size
                                docx_data = ole_data[pkg_offset:pkg_offset + pkg_size]
                                break
                elif ole_data[:4] == b'PK\x03\x04':
                    docx_data = ole_data

                if docx_data:
                    try:
                        import zipfile as _zf
                        pkg_zip = _zf.ZipFile(io.BytesIO(docx_data))
                        # 从document.xml提取第一个标题作为文件名
                        real_name = f'附件{ole_num}'
                        if 'word/document.xml' in pkg_zip.namelist():
                            doc_xml_inner = pkg_zip.read('word/document.xml').decode('utf-8')
                            texts = re.findall(r'<w:t[^>]*>([^<]+)</w:t>', doc_xml_inner)
                            if texts and len(texts[0]) > 1:
                                real_name = texts[0].strip()

                        # 递归转换：走完整docx_to_md流程（含嵌套OLE提取）
                        md_content, image_idx = _convert_docx_bytes_to_md(
                            docx_data, images_dir, attachments_dir, image_idx
                        )

                        ole_info[image_name] = {
                            'original_name': f'{real_name}.docx',
                            'saved_name': f'{real_name}.docx',
                            'md_name': f'{real_name}.md',
                            'md_content': md_content,
                            'is_image': False,
                            'has_md': True
                        }
                        print(f"Word嵌入对象提取(递归): {real_name}.docx")
                        continue
                    except Exception as e:
                        print(f"Word嵌入处理失败(附件{ole_num}): {e}")
                        continue
            elif prog_id == 'Visio.Drawing.11':
                embedded_data = ole_data
                original_name = f'附件{ole_num}.vsd'
                is_image = False
            elif prog_id == 'PBrush':
                bmp_start = ole_data.find(b'BM')
                if bmp_start != -1:
                    bmp_data = ole_data[bmp_start:]
                    try:
                        img = Image.open(io.BytesIO(bmp_data))
                        png_buffer = io.BytesIO()
                        img.save(png_buffer, format='PNG')
                        embedded_data = png_buffer.getvalue()
                        original_name = f'图片{ole_num}.png'
                        is_image = True
                    except:
                        continue
                else:
                    continue

            if original_name and embedded_data and image_name:
                # 清理文件名中的特殊字符，用于保存
                safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", original_name)

                # 保存到对应目录（按需创建）
                if is_image:
                    # 图片统一编号放入images文件夹，使用简短命名
                    images_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
                    numbered_name = f"image_{image_idx}.png"
                    save_path = images_dir / numbered_name
                    save_path.write_bytes(embedded_data)
                    print(f"提取图片: {original_name} -> {numbered_name}")
                    ole_info[image_name] = {
                        'original_name': original_name,
                        'saved_name': numbered_name,
                        'is_image': True,
                        'image_idx': image_idx
                    }
                    image_idx += 1
                else:
                    attachments_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
                    save_path = attachments_dir / safe_name
                    save_path.write_bytes(embedded_data)
                    print(f"提取附件: {original_name}")

                    md_name = None
                    name_lower = original_name.lower()

                    # 对doc/docx附件自动转为md（递归提取嵌套附件）
                    if name_lower.endswith('.docx'):
                        try:
                            md_content, image_idx = _convert_docx_bytes_to_md(
                                embedded_data, images_dir, attachments_dir, image_idx
                            )
                            md_name = f'{Path(original_name).stem}.md'
                            (attachments_dir / md_name).write_text(md_content, encoding='utf-8')
                            print(f"生成文本(md): {md_name}")
                        except Exception as e:
                            print(f"附件docx转md失败({original_name}): {e}")
                    elif name_lower.endswith('.doc'):
                        try:
                            md_content = _convert_doc_bytes_to_md(embedded_data, original_name)
                            md_name = f'{Path(original_name).stem}.md'
                            (attachments_dir / md_name).write_text(md_content, encoding='utf-8')
                            print(f"生成文本(md): {md_name}")
                        except Exception as e:
                            print(f"附件doc转md失败({original_name}): {e}")

                    ole_info[image_name] = {
                        'original_name': original_name,
                        'saved_name': safe_name,
                        'is_image': False,
                    }
                    if md_name:
                        ole_info[image_name]['md_name'] = md_name
                        ole_info[image_name]['has_md'] = True

    # 处理Excel的md文件生成（在ole_info中检查是否有Excel特殊记录）
    for image_name, info in list(ole_info.items()):
        if 'xlsx_data' in info:  # 这是Excel的特殊记录
            attachments_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
            # 保存xlsx文件
            original_xlsx = info['xlsx_name']
            xlsx_safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", original_xlsx)
            xlsx_path = attachments_dir / xlsx_safe_name
            xlsx_path.write_bytes(info['xlsx_data'])
            print(f"提取附件(xlsx): {original_xlsx} -> {xlsx_safe_name}")

            # 保存md文件
            original_md = info['md_name']
            md_safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", original_md)
            md_path = attachments_dir / md_safe_name
            md_path.write_text(info['md_content'], encoding='utf-8')
            print(f"生成文本(md): {original_md} -> {md_safe_name}")

            # 更新ole_info为标准格式
            ole_info[image_name] = {
                'original_name': original_xlsx,
                'saved_name': xlsx_safe_name,
                'md_name': md_safe_name,
                'is_image': False,
                'has_md': True
            }

    # 处理Word docx嵌入对象的保存
    for image_name, info in list(ole_info.items()):
        if 'md_content' in info and 'xlsx_data' not in info:
            attachments_dir.mkdir(parents=True, exist_ok=True)
            original_name = info['original_name']
            saved_name = info['saved_name']
            md_name = info['md_name']

            # 保存docx原文件
            att_path = attachments_dir / saved_name
            # docx数据在ole_data中，但这里只有md_content
            # 需要从ole重新获取...不过对于Word.Document.12，docx数据已经在前面处理时丢失
            # 这里只保存md即可，因为docx原数据是OLE容器格式不是标准docx

            # 保存md文件
            md_safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", md_name)
            md_path = attachments_dir / md_safe_name
            md_path.write_text(info['md_content'], encoding='utf-8')
            print(f"生成文本(md): {md_name} -> {md_safe_name}")

            # 更新ole_info
            ole_info[image_name] = {
                'original_name': original_name,
                'saved_name': saved_name,
                'md_name': md_safe_name,
                'is_image': False,
                'has_md': True
            }

    return ole_info, image_idx


def _xlrd_sheet_to_md(ws) -> str:
    """将xlrd worksheet转为markdown表格，自动裁剪空行空列"""
    all_rows = []
    for row_idx in range(ws.nrows):
        row = []
        for col_idx in range(ws.ncols):
            cell = ws.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                val = cell.value
                if val == int(val):
                    row.append(int(val))
                else:
                    row.append(val)
            else:
                row.append(cell.value)
        all_rows.append(tuple(row))

    # 过滤全空行
    all_rows = [row for row in all_rows if any(cell is not None for cell in row)]
    if not all_rows:
        return ""

    # 裁剪空的首列
    while all_rows and all(row and row[0] is None for row in all_rows):
        all_rows = [row[1:] for row in all_rows if len(row) > 1]
        if not any(all_rows):
            return ""

    # 裁剪空的尾列
    while all_rows and all(row and row[-1] is None for row in all_rows):
        all_rows = [row[:-1] for row in all_rows if len(row) > 1]
        if not any(all_rows):
            return ""

    result = ""
    for row_idx, row in enumerate(all_rows):
        if not row:
            continue
        cells = [str(cell).replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ') if cell is not None else '' for cell in row]
        result += "| " + " | ".join(cells) + " |\n"
        if row_idx == 0:
            result += "| " + " | ".join(["---"] * len(cells)) + " |\n"
    return result


def _worksheet_to_md(ws) -> str:
    """将openpyxl worksheet转为markdown表格，自动裁剪空行空列"""
    all_rows = list(ws.iter_rows(values_only=True))

    # 过滤全空行
    all_rows = [row for row in all_rows if any(cell is not None for cell in row)]
    if not all_rows:
        return ""

    # 裁剪空的首列
    while all_rows and all(row and row[0] is None for row in all_rows):
        all_rows = [row[1:] for row in all_rows if len(row) > 1]
        if not any(all_rows):
            return ""

    # 裁剪空的尾列
    while all_rows and all(row and row[-1] is None for row in all_rows):
        all_rows = [row[:-1] for row in all_rows if len(row) > 1]
        if not any(all_rows):
            return ""

    result = ""
    for row_idx, row in enumerate(all_rows):
        if not row:
            continue
        cells = [str(cell).replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ') if cell is not None else '' for cell in row]
        result += "| " + " | ".join(cells) + " |\n"
        if row_idx == 0:
            result += "| " + " | ".join(["---"] * len(cells)) + " |\n"
    return result


def xlsx_to_md(xlsx_path: str, output_dir: str = None) -> str:
    """
    将xlsx/xls文件转换为markdown格式

    Args:
        xlsx_path: xlsx/xls文件路径
        output_dir: 输出目录，默认为xlsx文件所在目录

    Returns:
        生成的md文件路径
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    if output_dir is None:
        output_dir = xlsx_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    base_name = xlsx_path.stem
    md_path = output_dir / f"{base_name}.md"

    print(f"处理Excel文件: {xlsx_path}")

    try:
        suffix = xlsx_path.suffix.lower()

        if suffix == '.xls':
            # 旧版.xls格式，用xlrd读取
            wb = xlrd.open_workbook(str(xlsx_path))
            sheet_names = wb.sheet_names()

            md_content = ""
            for sheet_name in sheet_names:
                ws = wb.sheet_by_name(sheet_name)
                md_content += f"## {sheet_name}\n\n"
                md_content += _xlrd_sheet_to_md(ws)
                md_content += "\n"

            md_path.write_text(md_content, encoding="utf-8")
            print(f"转换完成!")
            print(f"MD文件: {md_path}")
            print(f"包含 {len(sheet_names)} 个工作表")

        else:
            # 新版.xlsx格式，用openpyxl读取
            wb = load_workbook(xlsx_path, data_only=True)
            sheet_names = wb.sheetnames

            md_content = ""
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                md_content += f"## {sheet_name}\n\n"
                md_content += _worksheet_to_md(ws)
                md_content += "\n"

            wb.close()

            md_path.write_text(md_content, encoding="utf-8")
            print(f"转换完成!")
            print(f"MD文件: {md_path}")
            print(f"包含 {len(sheet_names)} 个工作表")

        return str(md_path)

    except Exception as e:
        raise RuntimeError(f"Excel转换失败: {e}")


def pdf_to_md(pdf_path: str, output_dir: str = None) -> str:
    """
    将PDF文件转换为markdown格式，使用MarkItDown提取文本 + 正则后处理识别结构

    Args:
        pdf_path: PDF文件路径
        output_dir: 输出目录，默认为PDF文件所在目录

    Returns:
        生成的md文件路径
    """
    from markitdown import MarkItDown

    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"文件不存在: {pdf_path}")

    if output_dir is None:
        output_dir = pdf_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    base_name = pdf_path.stem
    md_path = output_dir / f"{base_name}.md"

    print(f"处理PDF文件: {pdf_path}")

    # MarkItDown提取文本
    m = MarkItDown()
    result = m.convert(str(pdf_path))
    text = result.text_content

    # 正则后处理：识别文档结构，转为markdown标题
    lines = text.split('\n')
    new_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            new_lines.append('')
            continue

        # 第X章 -> #
        if re.match(r'^第[一二三四五六七八九十百零\d]+章', stripped):
            new_lines.append(f'# {stripped}')
        # 第X条 -> ##
        elif re.match(r'^第[一二三四五六七八九十百零\d]+条', stripped):
            new_lines.append(f'## {stripped}')
        # 一、二、三、...的短子项 -> ###
        elif re.match(r'^[一二三四五六七八九十]+、', stripped) and len(stripped) < 60:
            new_lines.append(f'### {stripped}')
        # （一）（二）...的短子项 -> ####
        elif re.match(r'^（[一二三四五六七八九十]+）', stripped) and len(stripped) < 60:
            new_lines.append(f'#### {stripped}')
        else:
            new_lines.append(stripped)

    text = '\n'.join(new_lines)

    # 去掉孤立的页码行（单独的数字）
    text = re.sub(r'\n\d+\n', '\n', text)
    # 清理多余空行
    while '\n\n\n' in text:
        text = text.replace('\n\n\n', '\n\n')

    md_path.write_text(text, encoding="utf-8")

    print(f"转换完成!")
    print(f"MD文件: {md_path}")

    return str(md_path)


def docx_to_md(docx_path: str, output_dir: str = None) -> str:
    """
    将docx文件转换为markdown格式，提取图片和嵌入附件

    Args:
        docx_path: docx文件路径
        output_dir: 输出目录，默认为docx文件所在目录

    Returns:
        生成的md文件路径
    """
    docx_path = Path(docx_path)
    if not docx_path.exists():
        raise FileNotFoundError(f"文件不存在: {docx_path}")

    if output_dir is None:
        output_dir = docx_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    base_name = docx_path.stem

    # 资源文件夹路径（按需创建）
    assets_dir = output_dir / f"{base_name}_files"
    images_dir = assets_dir / "images"
    attachments_dir = assets_dir / "attachments"

    # 先提取OLE嵌入对象，返回ole信息和下一个图片编号
    ole_info, next_image_idx = extract_ole_objects(docx_path, assets_dir, base_name)

    temp_media_dir = output_dir / "_temp_media"
    md_path = output_dir / f"{base_name}.md"

    cmd = [
        "pandoc",
        str(docx_path),
        "-f", "docx",
        "-t", "gfm",
        "--wrap=none",
        "--extract-media", str(temp_media_dir)
    ]

    print(f"执行命令: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, encoding="utf-8")

    if result.returncode != 0:
        print(f"pandoc错误: {result.stderr}")
        raise RuntimeError(f"pandoc转换失败: {result.stderr}")

    content = result.stdout

    # 修复pandoc转义的mermaid代码块
    def _unescape_mermaid(m):
        inner = m.group(1)
        inner = inner.replace('\\`', '`').replace('\\[', '[').replace('\\]', ']')
        inner = inner.replace('\\{', '{').replace('\\}', '}')
        inner = inner.replace('\\|', '|').replace('\\>', '>').replace('\\<', '<')
        inner = inner.replace('\\(', '(').replace('\\)', ')')
        inner = inner.replace('\\_', '_').replace('\\#', '#')
        inner = inner.replace('\\-', '-').replace('\\+', '+')
        inner = inner.replace('\\=', '=').replace('\\.', '.')
        inner = inner.replace('\\\\\n', '\n').replace('\\\n', '\n')
        inner = inner.replace('\\\\', '\\')
        return f'```mermaid\n{inner}\n```'

    content = re.sub(r'\\`\\`\\`mermaid\\?\n(.*?)\\`\\`\\`', _unescape_mermaid, content, flags=re.DOTALL)

    extracted_media_path = temp_media_dir / "media"

    # 初始化变量（防止分支未执行时报错）
    image_map = {}
    ole_replacements = {}
    idx = next_image_idx

    if extracted_media_path.exists():
        image_files = sorted(extracted_media_path.iterdir())
        idx = next_image_idx  # 继续OLE图片的编号

        for img_file in image_files:
            if img_file.is_file():
                # 检查是否是OLE对象的预览图
                if img_file.suffix.lower() == '.emf' and img_file.name in ole_info:
                    # OLE对象处理在后面统一进行
                    continue

                # 普通图片，使用简短命名（image_1.png而非长文件名前缀）
                images_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
                # EMF格式转为PNG（Markdown/Obsidian不支持EMF预览）
                if img_file.suffix.lower() == '.emf':
                    try:
                        from PIL import Image
                        img = Image.open(img_file)
                        new_name = f"image_{idx}.png"
                        new_path = images_dir / new_name
                        img.save(new_path, format='PNG')
                        image_map[img_file.name] = new_name
                        print(f"EMF转PNG: {img_file.name} -> {new_name}")
                    except Exception as e:
                        print(f"EMF转PNG失败({img_file.name}): {e}")
                        new_name = f"image_{idx}{img_file.suffix}"
                        new_path = images_dir / new_name
                        shutil.copy2(img_file, new_path)
                        image_map[img_file.name] = new_name
                else:
                    new_name = f"image_{idx}{img_file.suffix}"
                    new_path = images_dir / new_name
                    shutil.copy2(img_file, new_path)
                    image_map[img_file.name] = new_name
                    print(f"图片重命名: {img_file.name} -> {new_name}")
                idx += 1

        # 处理OLE对象（图片和附件），需要处理同一行多个OLE的情况
        for img_file in image_files:
            if img_file.is_file() and img_file.suffix.lower() == '.emf' and img_file.name in ole_info:
                info = ole_info[img_file.name]
                original_name = info['original_name']
                saved_name = info['saved_name']

                if info['is_image']:
                    # 图片，获取文件名（不带后缀）作为标题
                    name_without_ext = Path(original_name).stem
                    # 先用一个临时标记替换，后续统一处理格式
                    placeholder = f'__OLE_IMG_{img_file.name}__'
                    content = re.sub(
                        rf'!\[([^\]]*)\]\([^)]*{re.escape(img_file.name)}\)',
                        placeholder,
                        content
                    )
                    # 记录占位符对应的实际内容
                    ole_replacements[placeholder] = {
                        'type': 'image',
                        'title': name_without_ext,
                        'path': f'{base_name}_files/images/{saved_name}'
                    }
                    print(f"OLE图片标记: {original_name}")
                else:
                    # 附件
                    placeholder = f'__OLE_ATT_{img_file.name}__'
                    content = re.sub(
                        rf'!\[([^\]]*)\]\([^)]*{re.escape(img_file.name)}\)',
                        placeholder,
                        content
                    )
                    if info.get('has_md'):
                        md_name = info.get('md_name')
                        md_stem = Path(md_name).stem
                        # Obsidian双链：有md则链向md文本（带路径）
                        ole_replacements[placeholder] = {
                            'type': 'attachment',
                            'link': f'[[{base_name}_files/attachments/{md_stem}]]'
                        }
                    else:
                        # 无md的附件（如Visio），不生成双链，直接移除占位符
                        ole_replacements[placeholder] = {
                            'type': 'attachment',
                            'link': ''
                        }
                    print(f"OLE附件标记: {original_name}")

        shutil.rmtree(temp_media_dir)

        # 替换普通图片路径
        for old_name, new_name in image_map.items():
            content = re.sub(
                rf'<img[^>]*src="[^"]*{re.escape(old_name)}"[^>]*>',
                rf'![]({base_name}_files/images/{new_name})',
                content
            )
            content = re.sub(
                rf'!\[([^\]]*)\]\([^)]*{re.escape(old_name)}\)',
                rf'![\1]({base_name}_files/images/{new_name})',
                content
            )

        print(f"已更新md文件中的图片引用")

    # 处理HTML表格中的markdown图片，转换为<img>标签
    # 因为某些markdown渲染器在HTML表格中无法正确解析markdown图片语法

    # 处理 <td>...</td> 中直接包含的markdown图片（单个或多个）
    def replace_md_img_in_td(match):
        td_content = match.group(1)
        # 替换所有markdown图片为img标签
        td_content = re.sub(
            r'!\[\]\(([^)]+)\)',
            r'<img src="\1">',
            td_content
        )
        td_content = re.sub(
            r'!\[([^\]]*)\]\(([^)]+)\)',
            r'<img src="\2" alt="\1">',
            td_content
        )
        return f'<td>{td_content}</td>'

    content = re.sub(
        r'<td[^>]*>(.*?)</td>',
        replace_md_img_in_td,
        content,
        flags=re.DOTALL
    )

    # 处理OLE占位符，逐行处理同一行多个OLE的情况
    lines = content.split('\n')
    result_lines = []
    for line in lines:
        # 检查这一行是否包含OLE占位符
        has_ole = any(placeholder in line for placeholder in ole_replacements)

        if has_ole:
            # 找出这一行所有OLE占位符的顺序
            placeholders_in_line = []
            for placeholder in ole_replacements:
                if placeholder in line:
                    # 找到位置并记录顺序
                    pos = line.find(placeholder)
                    placeholders_in_line.append((pos, placeholder))
            # 按位置排序
            placeholders_in_line.sort(key=lambda x: x[0])

            # 构建新内容，每个OLE都单独一行
            # 先保留行前面的内容
            prefix = line[:placeholders_in_line[0][0]].strip()

            new_parts = []
            if prefix:
                new_parts.append(prefix)

            for pos, placeholder in placeholders_in_line:
                info = ole_replacements[placeholder]
                if info['type'] == 'image':
                    # 图片：标题 + 图片嵌入
                    new_parts.append(f"**{info['title']}**")
                    new_parts.append(f"![]({info['path']})")
                else:
                    # 附件wikilink：前后空行，确保Obsidian能正确识别
                    new_parts.append(f"\n{info['link']}\n")

            # 检查占位符后面是否还有内容
            last_placeholder_end = 0
            for pos, placeholder in placeholders_in_line:
                last_placeholder_end = max(last_placeholder_end, pos + len(placeholder))

            suffix = line[last_placeholder_end:].strip()
            if suffix:
                new_parts.append(suffix)

            # 每个部分单独一行
            line = '\n'.join(new_parts)

        result_lines.append(line)

    content = '\n'.join(result_lines)

    # 处理表格空表头
    lines = content.split('\n')
    new_lines = []
    i = 0
    skip_next = False

    while i < len(lines):
        if skip_next:
            skip_next = False
            i += 1
            continue

        line = lines[i]
        if re.match(r'^\|[\s:|-]+\|$', line.strip()):
            if new_lines:
                prev_line = new_lines[-1]
                if re.match(r'^\|[\s|]+\|$', prev_line.strip()):
                    cells = [c.strip() for c in prev_line.split('|') if c.strip()]
                    if not cells:
                        if i + 1 < len(lines):
                            next_line = lines[i + 1]
                            next_cells = [c.strip() for c in next_line.split('|') if c.strip()]
                            if next_cells:
                                new_lines.pop()
                                new_lines.append(next_line)
                                new_lines.append(line)
                                skip_next = True
                                i += 1
                                continue
        new_lines.append(line)
        i += 1

    content = '\n'.join(new_lines)

    # 为标题添加序号
    # 根据层级生成序号: # -> 1., ## -> 1.1, ### -> 1.1.1, #### -> 1.1.1.1
    lines = content.split('\n')
    result_lines = []
    counters = [0, 0, 0, 0, 0]  # 5级标题计数器

    for line in lines:
        # 检测标题层级
        heading_match = re.match(r'^(#{1,6})\s+(.*)$', line)
        if heading_match:
            level = len(heading_match.group(1))
            title_text = heading_match.group(2)

            # 更新计数器
            counters[level - 1] += 1
            # 重置下级计数器
            for i in range(level, 5):
                counters[i] = 0

            # 生成序号（中式格式：1、1.1、1.1.1）
            nums = [str(counters[j]) for j in range(level) if counters[j] > 0]
            if nums:
                number = '.'.join(nums) + ' '
                # 如果标题已有序号开头，不重复添加
                if not re.match(r'^\d+', title_text):
                    line = heading_match.group(1) + ' ' + number + title_text

        result_lines.append(line)

    content = '\n'.join(result_lines)

    # 处理列表中的HTML注释分隔符 <!-- -->
    # pandoc在父列表项后用<!-- -->分隔，子列表项需要缩进
    lines = content.split('\n')
    result_lines = []

    i = 0
    indent_next_items = False  # 标记后续列表项是否需要缩进

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # 处理HTML注释行
        if stripped == '<!-- -->':
            # 回溯找最近的非空列表项
            for j in range(len(result_lines) - 1, max(-1, len(result_lines) - 5), -1):
                prev_stripped = result_lines[j].strip()
                if prev_stripped == '':
                    continue
                if prev_stripped.startswith('-') or re.match(r'^\d+\.', prev_stripped):
                    indent_next_items = True
                    break
                else:
                    break
            i += 1
            continue

        # 处理需要缩进的内容（只处理列表项，附件链接不缩进并去掉前导空格）
        if stripped.startswith('[📎'):
            # 附件链接行，去掉前导空格
            line = stripped
        elif stripped.startswith('-') or re.match(r'^\d+\.', stripped):
            if indent_next_items:
                # 添加缩进
                existing_indent = len(line) - len(line.lstrip())
                line = ' ' * (existing_indent + 4) + line.lstrip()
        else:
            # 其他非空行，恢复缩进标记
            if stripped:
                indent_next_items = False

        result_lines.append(line)
        i += 1

    content = '\n'.join(result_lines)

    # 处理同一行多个附件链接，换行显示
    lines = content.split('\n')
    result_lines = []
    for line in lines:
        # 统计这一行中附件链接的数量（以[📎开头的）
        # 找出所有[📎 xxx](path)模式
        attachment_links = re.findall(r'\[📎[^\]]*\]\([^)]+\)', line)
        if len(attachment_links) > 1:
            # 多个链接，每个链接换行显示
            # 先保留行前面的非链接内容
            prefix_match = re.match(r'^([^[]*)', line)
            prefix = prefix_match.group(1) if prefix_match else ''
            # 构建新行：前缀 + 每个链接换行
            new_content = prefix.rstrip()
            for link in attachment_links:
                # 检查链接后是否有纯文本链接
                plain_text_pattern = rf'{re.escape(link)}\s*\[📄 纯文本\]\([^)]+\)'
                plain_match = re.search(plain_text_pattern, line)
                if plain_match:
                    new_content += '\n' + plain_match.group(0)
                else:
                    new_content += '\n' + link
            line = new_content
        result_lines.append(line)

    content = '\n'.join(result_lines)

    # 更新目录链接格式：转为Obsidian兼容的heading link
    lines = content.split('\n')
    result_lines = []

    for line in lines:
        # 处理目录链接：[N 标题 [页码](#锚点)](#锚点) -> [[#N 标题]]
        toc_match = re.match(r'^\[([\d\.\s]+)([^\[]+)\s+\[[^\]]+\]\(#([^)]+)\)\]\(#([^)]+)\)', line)
        if toc_match:
            number_part = toc_match.group(1).strip()
            title_part = toc_match.group(2).strip()
            # Obsidian heading link
            line = f'[[#{number_part} {title_part}]]'
        # 也处理已简化但仍是锚点格式的链接：[N 标题](#锚点) -> [[#N 标题]]
        elif re.match(r'^\[[\d\.\s]+[^\]]+\]\(#[^)]+\)$', line.strip()):
            simple_match = re.match(r'^\[([^\]]+)\]\(#([^)]+)\)$', line.strip())
            if simple_match:
                link_text = simple_match.group(1).strip()
                line = f'[[#{link_text}]]'

        result_lines.append(line)

    content = '\n'.join(result_lines)

    md_path.write_text(content, encoding="utf-8")

    print(f"\n转换完成!")
    print(f"MD文件: {md_path}")

    # 只在存在资源目录时才显示
    has_images = images_dir.exists() and any(images_dir.iterdir())
    has_attachments = attachments_dir.exists() and any(attachments_dir.iterdir())

    if has_images or has_attachments:
        print(f"资源目录: {assets_dir}")
        if has_images:
            print(f"  - 图片: {images_dir}")
        if has_attachments:
            print(f"  - 附件: {attachments_dir}")

    return str(md_path)


def doc_to_docx(doc_path: str) -> str:
    """将旧格式.doc转换为.docx，使用Word COM自动化，返回临时docx路径"""
    import atexit
    import tempfile

    doc_path = str(Path(doc_path).resolve())
    tmp_dir = tempfile.mkdtemp(prefix="docx2md_")
    tmp_docx = str(Path(tmp_dir) / (Path(doc_path).stem + ".docx"))

    try:
        import win32com.client
    except ImportError:
        raise RuntimeError(
            "转换.doc需要pywin32库，请运行: pip install pywin32\n"
            "或者安装LibreOffice并在PATH中可用"
        )

    print(f"将.doc转换为.docx: {Path(doc_path).name}")
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False

    def cleanup():
        try:
            word.Quit()
        except Exception:
            pass
        try:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        except Exception:
            pass

    atexit.register(cleanup)

    try:
        doc = word.Documents.Open(doc_path)
        doc.SaveAs(tmp_docx, FileFormat=16)  # 16 = wdFormatXMLDocument (.docx)
        doc.Close()
    except Exception as e:
        cleanup()
        raise RuntimeError(f".doc转.docx失败: {e}")

    print(f"转换完成: {tmp_docx}")
    return tmp_docx


def main():
    import sys
    import argparse

    parser = argparse.ArgumentParser(
        description='将文档转换为AI可读的Markdown格式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
支持格式:
  - doc/docx: Word文档，提取图片和OLE附件
  - xlsx/xls: Excel文件，转换为markdown表格
  - pdf: PDF文档，提取文本、表格和图片

输出结构:
  docx: {filename}.md + {filename}_files/images/ + {filename}_files/attachments/
  xlsx: {filename}.md (包含所有工作表)

示例:
  python docx2md.py document.docx
  python docx2md.py document.doc
  python docx2md.py spreadsheet.xlsx
  python docx2md.py document.docx -o ./output
        '''
    )
    parser.add_argument('input_file', help='要转换的文件路径 (doc/docx/xlsx/xls)')
    parser.add_argument('-o', '--output', help='输出目录（默认为输入文件所在目录）')

    args = parser.parse_args()

    input_path = Path(args.input_file)
    suffix = input_path.suffix.lower()

    try:
        if suffix in ['.xlsx', '.xls']:
            md_path = xlsx_to_md(args.input_file, args.output)
        elif suffix == '.docx':
            md_path = docx_to_md(args.input_file, args.output)
        elif suffix == '.doc':
            docx_path = doc_to_docx(args.input_file)
            # .doc转换时，输出目录取原文件所在目录（除非用户指定了-o）
            doc_output = args.output if args.output else str(input_path.parent)
            try:
                md_path = docx_to_md(docx_path, doc_output)
            finally:
                # 清理临时docx和临时目录
                tmp_dir = str(Path(docx_path).parent)
                try:
                    os.remove(docx_path)
                    os.rmdir(tmp_dir)
                except Exception:
                    pass
        elif suffix == '.pdf':
            md_path = pdf_to_md(args.input_file, args.output)
        else:
            print(f"不支持的文件格式: {suffix}")
            print("支持的格式: .doc, .docx, .xlsx, .xls, .pdf")
            sys.exit(1)

        print(f"\n转换成功!")
        print(f"MD文件: {md_path}")
    except FileNotFoundError as e:
        print(f"文件不存在: {e}")
        sys.exit(1)
    except RuntimeError as e:
        print(f"转换失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()